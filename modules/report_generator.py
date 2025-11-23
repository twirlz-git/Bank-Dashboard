"""
modules/report_generator.py - Generate XLSX, PDF and JSON reports with chart support
"""

import logging
from typing import Dict, Any, Optional, List  # Add List here
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import tempfile
import os
import json

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generate XLSX, PDF and JSON format reports with chart visualization"""
    
    def __init__(self):
        """Initialize report generator"""
        try:
            from modules.chart_generator import ChartGenerator
            self.chart_generator = ChartGenerator()
            self.charts_enabled = True
        except ImportError:
            logger.warning("ChartGenerator not available, charts will be disabled")
            self.chart_generator = None
            self.charts_enabled = False
        
        # Check PDF library availability
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
            self.pdf_enabled = True
            logger.info("PDF generation enabled")
        except ImportError:
            self.pdf_enabled = False
            logger.warning("reportlab not installed. PDF generation disabled. Install: pip install reportlab")
    
    def _convert_value_for_excel(self, value: Any) -> Any:
        """
        Convert complex data types to Excel-compatible formats.
        
        Args:
            value: Value to convert
            
        Returns:
            Excel-compatible value (str, int, float, bool, or None)
        """
        # Handle None
        if value is None:
            return "Н/Д"
        
        # Handle dict - convert to readable string
        if isinstance(value, dict):
            try:
                # Format dict as key: value pairs
                items = [f"{k}: {v}" for k, v in value.items()]
                return "; ".join(items)
            except:
                return str(value)
        
        # Handle list/tuple - convert to comma-separated string
        if isinstance(value, (list, tuple)):
            try:
                return ", ".join(str(item) for item in value)
            except:
                return str(value)
        
        # Handle bool
        if isinstance(value, bool):
            return "Да" if value else "Нет"
        
        # Handle datetime
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y %H:%M")
        
        # Handle numeric types - return as is
        if isinstance(value, (int, float)):
            return value
        
        # Handle str - return as is
        if isinstance(value, str):
            return value
        
        # For any other type - convert to string
        try:
            return str(value)
        except:
            return "Н/Д"
    
    def generate_xlsx_comparison(self, comparison_data: Dict[str, Any]) -> BytesIO:
        """Generate XLSX file with comparison table and optional chart"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Сравнение"
        
        # Add title
        ws['A1'] = "Сравнение банковских продуктов"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Add timestamp
        ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:D2')
        
        # Add comparison table
        comparison_df = comparison_data.get("comparison_table")
        if comparison_df is not None:
            for r_idx, row in enumerate(comparison_df.values, start=4):
                for c_idx, value in enumerate(row, start=1):
                    # Convert value to Excel-compatible format
                    excel_value = self._convert_value_for_excel(value)
                    ws.cell(row=r_idx, column=c_idx, value=excel_value)
            
            # Add headers
            headers = comparison_df.columns
            for c_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=c_idx, value=str(header))
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add insights sheet
        insights_ws = wb.create_sheet("Анализ")
        insights_ws['A1'] = "Ключевые выводы"
        insights_ws['A1'].font = Font(size=12, bold=True)
        
        insights = comparison_data.get("insights", [])
        for idx, insight in enumerate(insights, start=2):
            insights_ws[f'A{idx}'] = str(insight)
        
        # Add recommendation
        recommendation = comparison_data.get("recommendation", "Недостаточно данных")
        insights_ws['A10'] = "Рекомендация:"
        insights_ws['A10'].font = Font(bold=True)
        insights_ws['A11'] = str(recommendation)
        
        # Add chart if available
        if self.charts_enabled and self.chart_generator:
            try:
                chart_ws = wb.create_sheet("График")
                chart_ws['A1'] = "Визуализация сравнения"
                chart_ws['A1'].font = Font(size=12, bold=True)
                
                # Generate comparison chart
                fig = self.chart_generator.generate_comparison_chart(comparison_data)
                
                # Save chart as image and embed
                img_path = self._save_chart_temp(fig)
                if img_path:
                    img = XLImage(img_path)
                    chart_ws.add_image(img, 'A3')
                    os.unlink(img_path)  # Clean up temp file
                    
            except Exception as e:
                logger.error(f"Failed to add comparison chart: {e}")
        
        # Auto-adjust column widths
        for sheet_name in wb.sheetnames:
            worksheet = wb[sheet_name]
            for col_idx, column in enumerate(worksheet.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def generate_pdf_comparison(self, comparison_data: Dict[str, Any]) -> BytesIO:
        """
        Generate PDF report with comparison table, insights, and charts.
        
        Args:
            comparison_data: Comparison data dict with table, insights, advantages, recommendation
            
        Returns:
            BytesIO with PDF content
        """
        if not self.pdf_enabled:
            logger.error("PDF generation not available. Install reportlab: pip install reportlab")
            raise RuntimeError("PDF generation requires reportlab library")
        
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        
        # Container for elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=8
        )
        
        # Title
        title = Paragraph("Сравнение банковских продуктов", title_style)
        elements.append(title)
        
        # Timestamp
        timestamp = Paragraph(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal'])
        elements.append(timestamp)
        elements.append(Spacer(1, 0.3*inch))
        
        # LLM indicator
        if comparison_data.get("llm_powered", False):
            llm_note = Paragraph("🤖 <b>Сравнение сгенерировано с помощью LLM</b>", styles['Normal'])
            elements.append(llm_note)
            elements.append(Spacer(1, 0.2*inch))
        
        # Comparison table
        comparison_df = comparison_data.get("comparison_table")
        if comparison_df is not None:
            heading = Paragraph("📋 Сравнительная таблица", heading_style)
            elements.append(heading)
            elements.append(Spacer(1, 0.1*inch))
            
            # Prepare table data
            table_data = [comparison_df.columns.tolist()]  # Headers
            for row in comparison_df.values:
                table_data.append([self._convert_value_for_excel(val) for val in row])
            
            # Create table
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Key insights
        insights = comparison_data.get("insights", [])
        if insights:
            heading = Paragraph("💡 Ключевые выводы", heading_style)
            elements.append(heading)
            elements.append(Spacer(1, 0.1*inch))
            
            for insight in insights:
                # Clean markdown formatting for PDF
                clean_insight = self._clean_markdown_for_pdf(str(insight))
                insight_para = Paragraph(f"• {clean_insight}", styles['Normal'])
                elements.append(insight_para)
                elements.append(Spacer(1, 0.1*inch))
            
            elements.append(Spacer(1, 0.2*inch))
        
        # Advantages
        sber_advantages = comparison_data.get("sber_advantages", [])
        competitor_advantages = comparison_data.get("competitor_advantages", [])
        
        if sber_advantages or competitor_advantages:
            elements.append(PageBreak())
            
            if sber_advantages:
                heading = Paragraph("✅ Преимущества Сбера", heading_style)
                elements.append(heading)
                elements.append(Spacer(1, 0.1*inch))
                
                for adv in sber_advantages:
                    clean_adv = self._clean_markdown_for_pdf(str(adv))
                    adv_para = Paragraph(clean_adv, styles['Normal'])
                    elements.append(adv_para)
                    elements.append(Spacer(1, 0.1*inch))
                
                elements.append(Spacer(1, 0.2*inch))
            
            if competitor_advantages:
                heading = Paragraph("⚡ Преимущества конкурента", heading_style)
                elements.append(heading)
                elements.append(Spacer(1, 0.1*inch))
                
                for adv in competitor_advantages:
                    clean_adv = self._clean_markdown_for_pdf(str(adv))
                    adv_para = Paragraph(clean_adv, styles['Normal'])
                    elements.append(adv_para)
                    elements.append(Spacer(1, 0.1*inch))
                
                elements.append(Spacer(1, 0.2*inch))
        
        # Recommendation
        recommendation = comparison_data.get("recommendation", "")
        if recommendation:
            heading = Paragraph("🎯 Рекомендация", heading_style)
            elements.append(heading)
            elements.append(Spacer(1, 0.1*inch))
            
            clean_rec = self._clean_markdown_for_pdf(str(recommendation))
            rec_para = Paragraph(clean_rec, styles['Normal'])
            elements.append(rec_para)
        
        # Add chart if available
        if self.charts_enabled and self.chart_generator:
            try:
                elements.append(PageBreak())
                heading = Paragraph("📊 Визуализация", heading_style)
                elements.append(heading)
                elements.append(Spacer(1, 0.2*inch))
                
                fig = self.chart_generator.generate_comparison_chart(comparison_data)
                img_path = self._save_chart_temp(fig)
                
                if img_path:
                    img = RLImage(img_path, width=5*inch, height=3.5*inch)
                    elements.append(img)
                    os.unlink(img_path)
            except Exception as e:
                logger.error(f"Failed to add chart to PDF: {e}")
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def generate_pdf_trends(self, trends_data: Dict[str, Any]) -> BytesIO:
        """
        Generate PDF report with trends analysis and charts.
        
        Args:
            trends_data: Trends data dict with timeline, analysis, summary
            
        Returns:
            BytesIO with PDF content
        """
        if not self.pdf_enabled:
            logger.error("PDF generation not available. Install reportlab: pip install reportlab")
            raise RuntimeError("PDF generation requires reportlab library")
        
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
        from reportlab.lib.enums import TA_CENTER
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=8
        )
        
        # Title
        title = Paragraph("Анализ трендов банковских продуктов", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Metadata
        bank = trends_data.get('bank', 'Н/Д')
        product_type = trends_data.get('product_type', 'Н/Д')
        start_date = trends_data.get('start_date', 'Н/Д')
        end_date = trends_data.get('end_date', 'Н/Д')
        
        info_para = Paragraph(
            f"<b>Банк:</b> {bank}<br/>"
            f"<b>Продукт:</b> {product_type}<br/>"
            f"<b>Период:</b> {start_date} — {end_date}<br/>"
            f"<b>Дата отчета:</b> {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles['Normal']
        )
        elements.append(info_para)
        elements.append(Spacer(1, 0.3*inch))
        
        # Data source indicator
        data_source = trends_data.get('data_source', 'unknown')
        confidence = trends_data.get('confidence', 0.5)
        
        source_text = {
            'mock': '⚠️ Используются демо-данные',
            'real_data_based': '✅ Данные из реальных источников',
            'web_search': '✅ Данные получены через web-search'
        }.get(data_source, f'📊 Источник: {data_source}')
        
        source_para = Paragraph(f"{source_text} | Достоверность: {confidence:.0%}", styles['Normal'])
        elements.append(source_para)
        elements.append(Spacer(1, 0.2*inch))
        
        # Summary
        summary = trends_data.get('summary', '')
        if summary:
            heading = Paragraph("📊 Сводка", heading_style)
            elements.append(heading)
            elements.append(Spacer(1, 0.1*inch))
            
            clean_summary = self._clean_markdown_for_pdf(summary)
            summary_para = Paragraph(clean_summary, styles['Normal'])
            elements.append(summary_para)
            elements.append(Spacer(1, 0.3*inch))
        
        # Timeline table
        timeline = trends_data.get('timeline', [])
        if timeline:
            elements.append(PageBreak())
            heading = Paragraph("📋 Таблица изменений", heading_style)
            elements.append(heading)
            elements.append(Spacer(1, 0.1*inch))
            
            # Prepare table data
            table_data = [["Дата", "Ставка (%)", "Причина", "Достоверность"]]
            for item in timeline:
                table_data.append([
                    item.get("date", "Н/Д"),
                    f"{item.get('rate', 0):.2f}",
                    item.get("reason", "Н/Д"),
                    f"{item.get('confidence', 0.5):.0%}"
                ])
            
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Statistics
        analysis = trends_data.get('analysis', {})
        if analysis and analysis.get('status') == 'success':
            elements.append(PageBreak())
            heading = Paragraph("📈 Статистика", heading_style)
            elements.append(heading)
            elements.append(Spacer(1, 0.1*inch))
            
            stats_text = f"""
            <b>Начальное значение:</b> {analysis.get('start_value', 0):.2f}%<br/>
            <b>Конечное значение:</b> {analysis.get('end_value', 0):.2f}%<br/>
            <b>Среднее значение:</b> {analysis.get('average_value', 0):.2f}%<br/>
            <b>Диапазон:</b> {analysis.get('min_value', 0):.2f}% — {analysis.get('max_value', 0):.2f}%<br/>
            <b>Общее изменение:</b> {analysis.get('total_change', 0):+.2f}% ({analysis.get('change_percentage', 0):+.1f}%)<br/>
            <b>Точек данных:</b> {analysis.get('data_points', 0)}<br/>
            <b>Точек изменения:</b> {analysis.get('change_points', 0)}
            """
            stats_para = Paragraph(stats_text, styles['Normal'])
            elements.append(stats_para)
            elements.append(Spacer(1, 0.2*inch))
        
        # Add charts if available
        if self.charts_enabled and self.chart_generator and timeline:
            try:
                elements.append(PageBreak())
                heading = Paragraph("📊 Визуализация трендов", heading_style)
                elements.append(heading)
                elements.append(Spacer(1, 0.2*inch))
                
                fig = self.chart_generator.generate_timeline_chart(timeline, f"Динамика {product_type} - {bank}")
                img_path = self._save_chart_temp(fig)
                
                if img_path:
                    img = RLImage(img_path, width=6*inch, height=3.5*inch)
                    elements.append(img)
                    os.unlink(img_path)
                    
            except Exception as e:
                logger.error(f"Failed to add chart to PDF: {e}")
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def _clean_markdown_for_pdf(self, text: str) -> str:
        """
        Clean markdown formatting for PDF rendering.
        Converts basic markdown to HTML-like tags for reportlab.
        """
        import re
        
        # Bold: **text** -> <b>text</b>
        text = re.sub(r'\*\*([^*]+)\*\*', r'<b>\1</b>', text)
        
        # Italic: *text* -> <i>text</i>
        text = re.sub(r'\*([^*]+)\*', r'<i>\1</i>', text)
        
        # Remove markdown bullets (•) - they're added manually
        text = text.replace('•', '')
        
        # Line breaks
        text = text.replace('\n', '<br/>')
        
        return text
    def generate_pdf_multibank(self, comparison_data: Dict[str, Any], banks: List[str]) -> BytesIO:
        """Generate PDF report for multi-bank comparison"""
        if not self.pdf_enabled:
            raise RuntimeError("PDF generation requires reportlab library")

        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
        from reportlab.lib.enums import TA_CENTER

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        
        # Styles
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#366092'), alignment=TA_CENTER, spaceAfter=12)
        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#366092'), spaceAfter=8)
        
        # Title
        elements.append(Paragraph("Мульти-банк сравнение продуктов", title_style))
        elements.append(Paragraph(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        if comparison_data.get("llm_powered", False):
            elements.append(Paragraph("🤖 <b>Сравнение сгенерировано с помощью LLM</b>", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))

        # Table
        df = comparison_data.get("comparison_table")
        if df is not None:
            elements.append(Paragraph("📋 Сравнительная таблица", heading_style))
            elements.append(Spacer(1, 0.1*inch))
            
            table_data = [df.columns.tolist()]
            for row in df.values:
                table_data.append([self._convert_value_for_excel(val) for val in row])
            
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))

        # Insights
        insights = comparison_data.get("insights", [])
        if insights:
            elements.append(Paragraph("💡 Ключевые выводы", heading_style))
            for insight in insights:
                clean_insight = self._clean_markdown_for_pdf(str(insight))
                elements.append(Paragraph(f"• {clean_insight}", styles['Normal']))
                elements.append(Spacer(1, 0.05*inch))
            elements.append(Spacer(1, 0.2*inch))

        # Advantages (Sber)
        sber_adv = comparison_data.get("sber_advantages", [])
        if sber_adv:
             elements.append(Paragraph("✅ Преимущества Сбера", heading_style))
             for adv in sber_adv:
                 clean_adv = self._clean_markdown_for_pdf(str(adv))
                 elements.append(Paragraph(clean_adv, styles['Normal']))
                 elements.append(Spacer(1, 0.05*inch))
             elements.append(Spacer(1, 0.2*inch))

        # Competitor Highlights
        highlights_dict = comparison_data.get("competitor_highlights", {})
        if highlights_dict and banks:
             elements.append(Paragraph("⚡ Сильные стороны конкурентов", heading_style))
             for bank in banks:
                 h_list = highlights_dict.get(bank, [])
                 if h_list:
                     elements.append(Paragraph(f"<b>{bank}</b>", styles['Normal']))
                     for h in h_list:
                         clean_h = self._clean_markdown_for_pdf(str(h))
                         elements.append(Paragraph(f"- {clean_h}", styles['Normal']))
                     elements.append(Spacer(1, 0.1*inch))

        # Recommendation
        rec = comparison_data.get("recommendation", "")
        if rec:
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph("🎯 Рекомендация", heading_style))
            elements.append(Paragraph(self._clean_markdown_for_pdf(str(rec)), styles['Normal']))

        # Chart
        if self.charts_enabled and self.chart_generator:
             try:
                 elements.append(PageBreak())
                 elements.append(Paragraph("📊 Визуализация", heading_style))
                 fig = self.chart_generator.generate_comparison_chart(comparison_data)
                 img_path = self._save_chart_temp(fig)
                 if img_path:
                     img = RLImage(img_path, width=5*inch, height=3.5*inch)
                     elements.append(img)
                     os.unlink(img_path)
             except Exception as e:
                 logger.error(f"Chart error: {e}")

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_xlsx_multibank(self, comparison_data: Dict[str, Any], banks: List[str]) -> BytesIO:
        """Generate XLSX file for multi-bank comparison"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводная таблица"

        # Title
        ws['A1'] = "Мульти-банк сравнение продуктов"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:E2')

        # Comparison Table
        comparison_df = comparison_data.get("comparison_table")
        if comparison_df is not None:
            # Headers
            headers = comparison_df.columns
            for c_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=c_idx, value=str(header))
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Data
            for r_idx, row in enumerate(comparison_df.values, start=5):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=self._convert_value_for_excel(value))

        # Analysis Sheet
        analysis_ws = wb.create_sheet("Детальный анализ")
        
        # Insights
        analysis_ws['A1'] = "Ключевые выводы"
        analysis_ws['A1'].font = Font(size=12, bold=True)
        current_row = 2
        for insight in comparison_data.get("insights", []):
            analysis_ws[f'A{current_row}'] = str(insight)
            current_row += 1
        
        current_row += 2
        
        # Sber Advantages
        analysis_ws[f'A{current_row}'] = "Преимущества Сбербанка"
        analysis_ws[f'A{current_row}'].font = Font(size=12, bold=True)
        current_row += 1
        for adv in comparison_data.get("sber_advantages", []):
            analysis_ws[f'A{current_row}'] = str(adv)
            current_row += 1
            
        current_row += 2

        # Competitor Highlights
        analysis_ws[f'A{current_row}'] = "Сильные стороны конкурентов"
        analysis_ws[f'A{current_row}'].font = Font(size=12, bold=True)
        current_row += 1
        
        highlights = comparison_data.get("competitor_highlights", {})
        for bank in banks:
            bank_highlights = highlights.get(bank, [])
            if bank_highlights:
                analysis_ws[f'A{current_row}'] = f"--- {bank} ---"
                analysis_ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                for h in bank_highlights:
                    analysis_ws[f'A{current_row}'] = str(h)
                    current_row += 1
                current_row += 1

        # Recommendation
        current_row += 1
        analysis_ws[f'A{current_row}'] = "Общая рекомендация"
        analysis_ws[f'A{current_row}'].font = Font(size=12, bold=True)
        current_row += 1
        analysis_ws[f'A{current_row}'] = comparison_data.get("recommendation", "")
        analysis_ws[f'A{current_row}'].alignment = Alignment(wrap_text=True)

        # Charts (if enabled)
        if self.charts_enabled and self.chart_generator:
            try:
                chart_ws = wb.create_sheet("Графики")
                fig = self.chart_generator.generate_comparison_chart(comparison_data)
                img_path = self._save_chart_temp(fig)
                if img_path:
                    img = XLImage(img_path)
                    chart_ws.add_image(img, 'A1')
                    os.unlink(img_path)
            except Exception as e:
                logger.error(f"Failed to add multibank chart: {e}")

        # Auto-size columns
        for sheet_name in wb.sheetnames:
            worksheet = wb[sheet_name]
            for col_idx, column in enumerate(worksheet.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_xlsx_trends(self, trends_data: Dict[str, Any]) -> BytesIO:
        """Generate XLSX file with trends analysis and charts"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Тренды"
        
        # Add title
        ws['A1'] = "Анализ трендов"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Add period info
        bank = trends_data.get('bank', 'Н/Д')
        product_type = trends_data.get('product_type', 'Н/Д')
        start_date = trends_data.get('start_date', 'Н/Д')
        end_date = trends_data.get('end_date', 'Н/Д')
        
        ws['A2'] = f"Банк: {bank}"
        ws['A3'] = f"Продукт: {product_type}"
        ws['A4'] = f"Период: {start_date} - {end_date}"
        
        # Add timeline table if available
        timeline = trends_data.get("timeline", [])
        if timeline:
            # Headers with better styling
            headers = ["Дата", "Ставка (%)", "Причина", "Достоверность", "Источник"]
            for c_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=6, column=c_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Data rows
            for r_idx, item in enumerate(timeline, start=7):
                ws[f'A{r_idx}'] = item.get("date", "Н/Д")
                ws[f'B{r_idx}'] = item.get("rate", "Н/Д")
                ws[f'C{r_idx}'] = item.get("reason", "Н/Д")
                ws[f'D{r_idx}'] = f"{item.get('confidence', 0.5):.0%}"
                ws[f'E{r_idx}'] = item.get("source", "Н/Д")
        
        # Add analysis sheet
        analysis = trends_data.get('analysis', {})
        if analysis and analysis.get('status') == 'success':
            analysis_ws = wb.create_sheet("Статистика")
            analysis_ws['A1'] = "Статистический анализ"
            analysis_ws['A1'].font = Font(size=12, bold=True)
            
            stats = [
                ("Начальное значение", f"{analysis.get('start_value', 0):.2f}%"),
                ("Конечное значение", f"{analysis.get('end_value', 0):.2f}%"),
                ("Среднее значение", f"{analysis.get('average_value', 0):.2f}%"),
                ("Минимум", f"{analysis.get('min_value', 0):.2f}%"),
                ("Максимум", f"{analysis.get('max_value', 0):.2f}%"),
                ("Общее изменение", f"{analysis.get('total_change', 0):+.2f}%"),
                ("Изменение (%)", f"{analysis.get('change_percentage', 0):+.1f}%"),
                ("Точек данных", str(analysis.get('data_points', 0))),
                ("Точек изменения", str(analysis.get('change_points', 0))),
                ("Средняя достоверность", f"{analysis.get('average_confidence', 0):.0%}")
            ]
            
            for idx, (label, value) in enumerate(stats, start=3):
                analysis_ws[f'A{idx}'] = label
                analysis_ws[f'A{idx}'].font = Font(bold=True)
                analysis_ws[f'B{idx}'] = value
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Выводы")
        summary_ws['A1'] = "Итоговый анализ"
        summary_ws['A1'].font = Font(size=12, bold=True)
        
        summary = trends_data.get('summary', 'Недостаточно данных')
        # Split summary by lines and add to cells
        summary_lines = summary.split('\n')
        for idx, line in enumerate(summary_lines, start=3):
            summary_ws[f'A{idx}'] = line
            summary_ws[f'A{idx}'].alignment = Alignment(wrap_text=True)
        
        # Add charts if available
        if self.charts_enabled and self.chart_generator and timeline:
            try:
                # Create charts sheet
                charts_ws = wb.create_sheet("Графики")
                charts_ws['A1'] = "Визуализация трендов"
                charts_ws['A1'].font = Font(size=12, bold=True)
                
                # Generate timeline chart
                fig1 = self.chart_generator.generate_timeline_chart(timeline, f"Динамика {product_type} - {bank}")
                img_path1 = self._save_chart_temp(fig1)
                if img_path1:
                    img1 = XLImage(img_path1)
                    img1.width = 800
                    img1.height = 400
                    charts_ws.add_image(img1, 'A3')
                    os.unlink(img_path1)
                
                # Generate detailed analysis chart if analysis available
                if analysis and analysis.get('status') == 'success':
                    fig2 = self.chart_generator.generate_trend_analysis_chart(timeline, analysis)
                    img_path2 = self._save_chart_temp(fig2)
                    if img_path2:
                        img2 = XLImage(img_path2)
                        img2.width = 800
                        img2.height = 600
                        charts_ws.add_image(img2, 'A28')  # Place below first chart
                        os.unlink(img_path2)
                        
            except Exception as e:
                logger.error(f"Failed to add trend charts: {e}")
        
        # Auto-adjust column widths
        for sheet_name in wb.sheetnames:
            worksheet = wb[sheet_name]
            for col_idx, column in enumerate(worksheet.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 80)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _save_chart_temp(self, fig) -> Optional[str]:
        """Save chart to temporary file and return path"""
        try:
            # Create temporary file
            temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            temp_path = temp_file.name
            temp_file.close()
            
            # Save chart as PNG
            fig.write_image(temp_path, format='png', width=800, height=500)
            
            return temp_path
        except Exception as e:
            logger.error(f"Failed to save chart to temp file: {e}")
            return None
    
    def get_filename(self, mode: str, bank: str = "", product_type: str = "", format: str = "xlsx") -> str:
        """Generate filename for report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if mode == "urgent":
            return f"сравнение_{bank}_{timestamp}.{format}"
        elif mode == "multibank":
            # Для мультибанка используем тип продукта в названии
            return f"мультибанк_{product_type}_{timestamp}.{format}"
        else:
            return f"тренды_{bank}_{product_type}_{timestamp}.{format}"

